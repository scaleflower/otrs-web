"""
OTRS Ticket Analysis Web Application
Flask-based ticket data analysis web application
"""

from flask import Flask, render_template, request, send_file, jsonify
from flask_sqlalchemy import SQLAlchemy
import pandas as pd
import numpy as np
from datetime import datetime
import re
import os
import io
import matplotlib.pyplot as plt
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///otrs_data.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Application version
APP_VERSION = "1.2.0"

# Initialize database
db = SQLAlchemy(app)

# Ensure upload directory exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Global variable to store uploaded data (deprecated - now using database)
# uploaded_data = {}

# Global variable to store processing progress
processing_status = {
    'current_step': 0,
    'total_steps': 7,
    'message': '',
    'details': ''
}

def update_processing_status(step, message, details=''):
    """Update processing status"""
    global processing_status
    processing_status['current_step'] = step
    processing_status['message'] = message
    processing_status['details'] = details
    print(f"Step {step}/{processing_status['total_steps']}: {message} - {details}")

def log_database_operation(operation_type, table_name, records_affected=0, operation_details='', filename=''):
    """Log database operations to log table"""
    try:
        # Get user information (IP address)
        user_ip = request.remote_addr if request else 'unknown'
        user_agent = request.headers.get('User-Agent', 'unknown') if request else 'unknown'
        user_info = f"IP: {user_ip}, Browser: {user_agent[:100]}"
        
        # Create log entry
        log_entry = DatabaseLog(
            operation_type=operation_type,
            table_name=table_name,
            records_affected=records_affected,
            operation_details=operation_details,
            user_info=user_info,
            filename=filename
        )
        
        db.session.add(log_entry)
        db.session.commit()
        print(f"Database operation logged: {operation_type} on {table_name}, affected {records_affected} records")
        
    except Exception as e:
        print(f"Error logging database operation: {str(e)}")
        db.session.rollback()

# Database models
class OtrsTicket(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    ticket_number = db.Column(db.String(100), unique=True, nullable=False)
    created_date = db.Column(db.DateTime)
    closed_date = db.Column(db.DateTime)
    state = db.Column(db.String(100))
    priority = db.Column(db.String(50))
    first_response = db.Column(db.String(255))
    age = db.Column(db.String(50))
    age_hours = db.Column(db.Float)
    
    # Other common fields from Excel
    queue = db.Column(db.String(255))
    owner = db.Column(db.String(255))
    customer_id = db.Column(db.String(255))
    customer_realname = db.Column(db.String(255))
    title = db.Column(db.Text)
    service = db.Column(db.String(255))
    type = db.Column(db.String(100))
    category = db.Column(db.String(255))
    sub_category = db.Column(db.String(255))
    responsible = db.Column(db.String(255))  # Responsible person field
    
    # Metadata
    import_time = db.Column(db.DateTime, default=datetime.utcnow)
    data_source = db.Column(db.String(255))  # Original filename
    raw_data = db.Column(db.Text)  # Store complete raw JSON data


# Upload record table
class UploadDetail(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(255), nullable=False)
    upload_time = db.Column(db.DateTime, default=datetime.utcnow)
    record_count = db.Column(db.Integer, nullable=False)
    import_mode = db.Column(db.String(50))  # Import mode: clear_existing or incremental


# Statistics query record table
class Statistic(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    query_time = db.Column(db.DateTime, default=datetime.utcnow)
    query_type = db.Column(db.String(50))  # Query type: main_analysis, age_details, empty_firstresponse, export_excel, export_txt
    total_records = db.Column(db.Integer)
    current_open_count = db.Column(db.Integer)
    empty_firstresponse_count = db.Column(db.Integer)
    daily_new_count = db.Column(db.Integer)  # Total new tickets count
    daily_closed_count = db.Column(db.Integer)  # Total closed tickets count
    age_segment = db.Column(db.String(50))  # Age segment (only for age_details queries)
    record_count = db.Column(db.Integer)  # Query result record count
    upload_id = db.Column(db.Integer, db.ForeignKey('upload_detail.id'))
    upload = db.relationship('UploadDetail', backref=db.backref('statistics', lazy=True))


# Database operation log table
class DatabaseLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    operation_time = db.Column(db.DateTime, default=datetime.utcnow)
    operation_type = db.Column(db.String(50))  # Operation type: clear_tickets, upload, delete, update, etc.
    table_name = db.Column(db.String(50))  # Table name
    records_affected = db.Column(db.Integer)  # Records affected
    operation_details = db.Column(db.Text)  # Operation details
    user_info = db.Column(db.String(255))  # User information (IP, browser, etc.)
    filename = db.Column(db.String(255))  # Related filename (if any)


# Responsible configuration table for storing user selections
class ResponsibleConfig(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_identifier = db.Column(db.String(255))  # User IP address for identification
    selected_responsibles = db.Column(db.Text)  # JSON array of selected responsible names
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

# Create database tables
with app.app_context():
    db.create_all()

def parse_age_to_hours(age_str):
    """Parse Age string to total hours"""
    if pd.isna(age_str):
        return 0
    
    age_str = str(age_str).lower()
    days = 0
    hours = 0
    minutes = 0
    
    day_match = re.search(r'(\d+)\s*d', age_str)
    if day_match:
        days = int(day_match.group(1))
    
    hour_match = re.search(r'(\d+)\s*h', age_str)
    if hour_match:
        hours = int(hour_match.group(1))
    
    minute_match = re.search(r'(\d+)\s*m', age_str)
    if minute_match:
        minutes = int(minute_match.group(1))
    
    return (days * 24) + hours + (minutes / 60)

def analyze_otrs_tickets_from_db():
    """Main function for OTRS ticket data analysis from database"""
    # Get all tickets from database
    tickets = OtrsTicket.query.all()
    
    if not tickets:
        return {}
    
    # Convert tickets to DataFrame for analysis
    ticket_data = []
    for ticket in tickets:
        ticket_data.append({
            'TicketNumber': ticket.ticket_number,
            'Created': ticket.created_date,
            'Closed': ticket.closed_date,
            'State': ticket.state,
            'Priority': ticket.priority,
            'FirstResponse': ticket.first_response,
            'Age': ticket.age,
            'AgeHours': ticket.age_hours
        })
    
    df = pd.DataFrame(ticket_data)
    
    # Execute ticket statistical analysis
    return analyze_ticket_statistics(df, {
        'ticket_number': 'TicketNumber',
        'created': 'Created',
        'closed': 'Closed',
        'state': 'State',
        'priority': 'Priority',
        'firstresponse': 'FirstResponse',
        'age': 'Age'
    })

def analyze_otrs_tickets_direct_from_db():
    """Main function for OTRS ticket data analysis directly from database using SQL queries"""
    stats = {}
    
    # Total records
    total_records = OtrsTicket.query.count()
    stats['total_records'] = total_records
    
    if total_records == 0:
        return stats
    
    # Current open tickets (where closed_date is NULL)
    current_open_count = OtrsTicket.query.filter(OtrsTicket.closed_date.is_(None)).count()
    stats['current_open_count'] = current_open_count
    
    # Empty first response (where first_response is NULL or empty, and state is not Closed/Resolved)
    empty_firstresponse_count = OtrsTicket.query.filter(
        (OtrsTicket.first_response.is_(None) | 
         (OtrsTicket.first_response == '') |
         (OtrsTicket.first_response == 'nan') |
         (OtrsTicket.first_response == 'NaN')),
        ~OtrsTicket.state.in_(['Closed', 'Resolved'])
    ).count()
    stats['empty_firstresponse_count'] = empty_firstresponse_count
    
    # Daily new tickets count
    daily_new = db.session.query(
        db.func.date(OtrsTicket.created_date).label('date'),
        db.func.count(OtrsTicket.id).label('count')
    ).filter(OtrsTicket.created_date.isnot(None)).group_by(db.func.date(OtrsTicket.created_date)).all()
    
    stats['daily_new'] = {str(record.date): record.count for record in daily_new}
    
    # Daily closed tickets count
    daily_closed = db.session.query(
        db.func.date(OtrsTicket.closed_date).label('date'),
        db.func.count(OtrsTicket.id).label('count')
    ).filter(OtrsTicket.closed_date.isnot(None)).group_by(db.func.date(OtrsTicket.closed_date)).all()
    
    stats['daily_closed'] = {str(record.date): record.count for record in daily_closed}
    
    # Calculate cumulative open tickets
    if stats['daily_new'] and stats['daily_closed']:
        daily_open = {}
        cumulative_open = 0
        
        # Get all dates and sort them
        all_dates = sorted(set(stats['daily_new'].keys()) | set(stats['daily_closed'].keys()))
        
        for date in all_dates:
            new_count = stats['daily_new'].get(date, 0)
            closed_count = stats['daily_closed'].get(date, 0)
            cumulative_open = cumulative_open + new_count - closed_count
            daily_open[date] = cumulative_open
        
        stats['daily_open'] = daily_open
    
    # Priority distribution
    priority_distribution = db.session.query(
        OtrsTicket.priority,
        db.func.count(OtrsTicket.id).label('count')
    ).filter(OtrsTicket.priority.isnot(None)).group_by(OtrsTicket.priority).all()
    
    stats['priority_distribution'] = {record.priority: record.count for record in priority_distribution}
    
    # State distribution
    state_distribution = db.session.query(
        OtrsTicket.state,
        db.func.count(OtrsTicket.id).label('count')
    ).filter(OtrsTicket.state.isnot(None)).group_by(OtrsTicket.state).all()
    
    stats['state_distribution'] = {record.state: record.count for record in state_distribution}
    
    # Age segments for open tickets
    open_tickets = OtrsTicket.query.filter(OtrsTicket.closed_date.is_(None)).all()
    
    age_segments = {
        'age_24h': 0,
        'age_24_48h': 0,
        'age_48_72h': 0,
        'age_72h': 0
    }
    
    for ticket in open_tickets:
        if ticket.age_hours is not None:
            if ticket.age_hours <= 24:
                age_segments['age_24h'] += 1
            elif ticket.age_hours <= 48:
                age_segments['age_24_48h'] += 1
            elif ticket.age_hours <= 72:
                age_segments['age_48_72h'] += 1
            else:
                age_segments['age_72h'] += 1
    
    stats['age_segments'] = age_segments
    
    return stats

def analyze_ticket_statistics(df, columns):
    """Perform ticket statistical analysis"""
    stats = {}
    
    # Convert date columns to datetime format
    if 'created' in columns:
        df['created_date'] = pd.to_datetime(df[columns['created']], errors='coerce')
        daily_new = df['created_date'].dt.date.value_counts().sort_index()
        # Convert date keys to string for JSON serialization
        stats['daily_new'] = {str(date): count for date, count in daily_new.items()}
    
    if 'closed' in columns:
        df['closed_date'] = pd.to_datetime(df[columns['closed']], errors='coerce')
        closed_tickets = df[df['closed_date'].notna()]
        if not closed_tickets.empty:
            daily_closed = closed_tickets['closed_date'].dt.date.value_counts().sort_index()
            # Convert date keys to string for JSON serialization
            stats['daily_closed'] = {str(date): count for date, count in daily_closed.items()}
            
            # Calculate cumulative open tickets if both daily_new and daily_closed exist
            if 'daily_new' in stats and 'daily_closed' in stats:
                daily_open = {}
                cumulative_open = 0
                
                # Get all dates and sort them
                all_dates = sorted(set(stats['daily_new'].keys()) | set(stats['daily_closed'].keys()))
                
                for date in all_dates:
                    new_count = stats['daily_new'].get(date, 0)
                    closed_count = stats['daily_closed'].get(date, 0)
                    cumulative_open = cumulative_open + new_count - closed_count
                    daily_open[date] = cumulative_open
                
                stats['daily_open'] = daily_open
    
    # Current open tickets analysis
    if 'closed' in columns:
        current_open = df[df['closed_date'].isna()]
        stats['current_open_count'] = len(current_open)
        
        # State distribution
        if 'state' in columns:
            state_counts = df[columns['state']].value_counts().to_dict()
            stats['state_distribution'] = state_counts
    
    # Priority distribution
    if 'priority' in columns:
        priority_counts = df[columns['priority']].value_counts().to_dict()
        stats['priority_distribution'] = priority_counts
    
        # FirstResponse empty analysis - exclude Closed and Resolved states
        if 'firstresponse' in columns:
            firstresponse_col = columns['firstresponse']
            nan_empty = df[firstresponse_col].isna()
            empty_strings = df[firstresponse_col] == ''
            nan_strings = df[firstresponse_col].astype(str).str.lower() == 'nan'
            
            # Exclude Closed and Resolved states
            if 'state' in columns:
                # Filter out Closed and Resolved states
                not_closed_resolved = ~df[columns['state']].isin(['Closed', 'Resolved'])
                empty_firstresponse = df[(nan_empty | empty_strings | nan_strings) & not_closed_resolved]
            else:
                empty_firstresponse = df[nan_empty | empty_strings | nan_strings]
                
            stats['empty_firstresponse_count'] = len(empty_firstresponse)
            
            if 'priority' in columns:
                priority_empty_counts = empty_firstresponse[columns['priority']].value_counts().to_dict()
                stats['empty_firstresponse_by_priority'] = priority_empty_counts
    
    # Age analysis if Age column exists
    if 'Age' in df.columns:
        df['age_hours'] = df['Age'].apply(parse_age_to_hours)
        
        # Age segment statistics (only for Open tickets)
        if 'closed' in columns:
            open_tickets = df[df['closed_date'].isna()]
            age_segments = {
                'age_24h': len(open_tickets[open_tickets['age_hours'] <= 24]),
                'age_24_48h': len(open_tickets[(open_tickets['age_hours'] > 24) & (open_tickets['age_hours'] <= 48)]),
                'age_48_72h': len(open_tickets[(open_tickets['age_hours'] > 48) & (open_tickets['age_hours'] <= 72)]),
                'age_72h': len(open_tickets[open_tickets['age_hours'] > 72])
            }
            stats['age_segments'] = age_segments
        
        stats['age_analysis'] = {
            'total_tickets': len(df),
            'tickets_with_age': len(df[df['Age'].notna()]),
            'avg_age_hours': df['age_hours'].mean() if not df['age_hours'].empty else 0
        }
    
    return stats

def generate_histogram(daily_new, daily_closed, daily_open=None):
    """Generate histogram for daily ticket statistics"""
    fig, ax = plt.subplots(figsize=(12, 6))
    
    # Prepare data
    all_dates = sorted(set(daily_new.keys()) | set(daily_closed.keys()))
    new_counts = [daily_new.get(date, 0) for date in all_dates]
    closed_counts = [daily_closed.get(date, 0) for date in all_dates]
    
    # Calculate open counts if provided
    if daily_open:
        open_counts = [daily_open.get(date, 0) for date in all_dates]
        bar_width = 0.25
        x_pos = np.arange(len(all_dates))
        
        ax.bar(x_pos - bar_width, new_counts, bar_width, label='New Tickets', alpha=0.8, color='#2ecc71')
        ax.bar(x_pos, closed_counts, bar_width, label='Closed Tickets', alpha=0.8, color='#e74c3c')
        ax.bar(x_pos + bar_width, open_counts, bar_width, label='Open Tickets', alpha=0.8, color='#3498db')
    else:
        bar_width = 0.4
        x_pos = np.arange(len(all_dates))
        
        ax.bar(x_pos - bar_width/2, new_counts, bar_width, label='New Tickets', alpha=0.8, color='#2ecc71')
        ax.bar(x_pos + bar_width/2, closed_counts, bar_width, label='Closed Tickets', alpha=0.8, color='#e74c3c')
    
    ax.set_xlabel('Date')
    ax.set_ylabel('Number of Tickets')
    ax.set_title('Daily Ticket Statistics')
    ax.set_xticks(x_pos)
    ax.set_xticklabels([str(date) for date in all_dates], rotation=45, ha='right')
    ax.legend()
    ax.grid(True, alpha=0.3)
    
    plt.tight_layout()
    
    # Save to buffer
    img_buffer = io.BytesIO()
    plt.savefig(img_buffer, format='png', dpi=100, bbox_inches='tight')
    img_buffer.seek(0)
    plt.close()
    
    return img_buffer

@app.route('/')
def index():
    """Main page"""
    return render_template('index.html', APP_VERSION=APP_VERSION)

@app.route('/uploads')
def view_uploads():
    """View all uploaded data sources"""
    data_sources = OtrsTicket.query.with_entities(OtrsTicket.data_source, db.func.count(OtrsTicket.id)).group_by(OtrsTicket.data_source).all()
    return render_template('uploads.html', data_sources=data_sources)

@app.route('/upload/<filename>')
def view_upload_details(filename):
    """View details of a specific upload file"""
    tickets = OtrsTicket.query.filter_by(data_source=filename).all()
    return render_template('upload_details.html', filename=filename, tickets=tickets)

@app.route('/upload', methods=['POST'])
def upload_file():
    """Handle file upload and analysis"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    # Get clear existing option from form data
    clear_existing = request.form.get('clear_existing', 'false').lower() == 'true'
    
    if file and file.filename.endswith(('.xlsx', '.xls')):
        try:
            # Step 1: Start processing file
            update_processing_status(1, 'Start processing Excel file', 'Reading file...')
            
            # Read Excel file
            df = pd.read_excel(file)
            update_processing_status(2, 'Excel file read completed', f'Found {len(df)} records in total')
            
            # Clear existing data if requested
            if clear_existing:
                update_processing_status(3, 'Clearing existing data', 'Deleting old records from database...')
                # Get current record count for logging
                existing_record_count = OtrsTicket.query.count()
                OtrsTicket.query.delete()
                print(f"Cleared existing data, importing {len(df)} new records")
                
                # Log database operation
                log_database_operation(
                    operation_type='clear_tickets',
                    table_name='otrs_ticket',
                    records_affected=existing_record_count,
                    operation_details='Cleared existing data when uploading file',
                    filename=file.filename
                )
            
            # Find actual column names with more comprehensive mapping
            update_processing_status(4, 'Analyzing Excel column structure', 'Identifying ticket data columns...')
            possible_columns = {
                'ticket_number': ['Ticket Number', 'TicketNumber', 'Number', 'ticket_number', 'id', 'Ticket', 'Ticket ID'],
                'created': ['Created', 'CreateTime', 'Create Time', 'Date Created', 'created', 'creation_date', 'Create Date'],
                'closed': ['Closed', 'CloseTime', 'Close Time', 'Date Closed', 'closed', 'close_date', 'Close Date'],
                'state': ['State', 'Status', 'Ticket State', 'state', 'status', 'Ticket Status'],
                'priority': ['Priority', 'priority', 'Ticket Priority'],
                'firstresponse': ['FirstResponse', 'First Response', 'firstresponse', 'First Reply', 'First Reply Time'],
                'age': ['Age', 'age', 'Ticket Age', 'Age of Ticket'],
                'queue': ['Queue', 'queue', 'Ticket Queue'],
                'owner': ['Owner', 'owner', 'Ticket Owner', 'Assigned To'],
                'customer_id': ['CustomerID', 'Customer ID', 'customer_id', 'Customer'],
                'customer_realname': ['Customer Realname', 'Customer Name', 'Customer Real Name'],
                'title': ['Title', 'title', 'Ticket Title', 'Subject'],
                'service': ['Service', 'service', 'Ticket Service'],
                'type': ['Type', 'type', 'Ticket Type'],
                'category': ['Category', 'category', 'Ticket Category'],
                'sub_category': ['Sub Category', 'SubCategory', 'sub_category', 'Ticket Sub Category'],
                'responsible': ['Responsible', 'responsible', 'Assignee', 'assignee', '处理人', '负责人']
            }
            
            actual_columns = {}
            for key, possible_names in possible_columns.items():
                for col in df.columns:
                    if any(name.lower() in col.lower() for name in possible_names):
                        actual_columns[key] = col
                        break
            
            # Save each ticket to database
            update_processing_status(5, 'Importing data to database', 'Saving ticket records...')
            new_records_count = 0
            total_records = len(df)
            
            for index, (_, row) in enumerate(df.iterrows()):
                # Parse dates
                created_date = None
                closed_date = None
                
                if 'created' in actual_columns:
                    try:
                        created_date = pd.to_datetime(row[actual_columns['created']], errors='coerce')
                        if pd.isna(created_date):
                            created_date = None
                    except:
                        created_date = None
                
                if 'closed' in actual_columns:
                    try:
                        closed_date = pd.to_datetime(row[actual_columns['closed']], errors='coerce')
                        if pd.isna(closed_date):
                            closed_date = None
                    except:
                        closed_date = None
                
                # Parse age to hours
                age_hours = 0
                if 'age' in actual_columns:
                    age_hours = parse_age_to_hours(row[actual_columns['age']])
                
                # Check if ticket already exists (for incremental import)
                ticket_number = str(row[actual_columns.get('ticket_number', '')]) if 'ticket_number' in actual_columns else None
                
                if not clear_existing and ticket_number:
                    existing_ticket = OtrsTicket.query.filter_by(ticket_number=ticket_number).first()
                    if existing_ticket:
                        continue  # Skip existing tickets in incremental mode
                
                # Create ticket record
                ticket = OtrsTicket(
                    ticket_number=ticket_number,
                    created_date=created_date,
                    closed_date=closed_date,
                    state=str(row[actual_columns.get('state', '')]) if 'state' in actual_columns else None,
                    priority=str(row[actual_columns.get('priority', '')]) if 'priority' in actual_columns else None,
                    first_response=str(row[actual_columns.get('firstresponse', '')]) if 'firstresponse' in actual_columns else None,
                    age=str(row[actual_columns.get('age', '')]) if 'age' in actual_columns else None,
                    age_hours=age_hours,
                    queue=str(row[actual_columns.get('queue', '')]) if 'queue' in actual_columns else None,
                    owner=str(row[actual_columns.get('owner', '')]) if 'owner' in actual_columns else None,
                    customer_id=str(row[actual_columns.get('customer_id', '')]) if 'customer_id' in actual_columns else None,
                    customer_realname=str(row[actual_columns.get('customer_realname', '')]) if 'customer_realname' in actual_columns else None,
                    title=str(row[actual_columns.get('title', '')]) if 'title' in actual_columns else None,
                    service=str(row[actual_columns.get('service', '')]) if 'service' in actual_columns else None,
                    type=str(row[actual_columns.get('type', '')]) if 'type' in actual_columns else None,
                    category=str(row[actual_columns.get('category', '')]) if 'category' in actual_columns else None,
                    sub_category=str(row[actual_columns.get('sub_category', '')]) if 'sub_category' in actual_columns else None,
                    responsible=str(row[actual_columns.get('responsible', '')]) if 'responsible' in actual_columns else None,
                    data_source=file.filename,
                    raw_data=row.to_json()
                )
                db.session.add(ticket)
                new_records_count += 1
                
                # Update progress every 100 records
                if index % 100 == 0:
                    update_processing_status(5, 'Importing data to database', 
                                           f'Processed {index + 1}/{total_records} records ({int((index + 1) / total_records * 100)}%)')
            
            # Commit all database changes
            update_processing_status(6, 'Committing database changes', 'Saving all data...')
            db.session.commit()
            
            # Record upload details to upload_detail table
            import_mode = 'clear_existing' if clear_existing else 'incremental'
            upload_record = UploadDetail(
                filename=file.filename,
                record_count=new_records_count,
                import_mode=import_mode
            )
            db.session.add(upload_record)
            db.session.commit()
            
            # Perform analysis directly from database using SQL queries
            update_processing_status(7, 'Generating statistical analysis', 'Calculating statistical metrics...')
            stats = analyze_otrs_tickets_direct_from_db()
            
            # Record statistical query results to statistic table
            total_new = sum(stats.get('daily_new', {}).values()) if 'daily_new' in stats else 0
            total_closed = sum(stats.get('daily_closed', {}).values()) if 'daily_closed' in stats else 0
            
            statistic_record = Statistic(
                query_type='main_analysis',
                total_records=len(df),
                current_open_count=stats.get('current_open_count', 0),
                empty_firstresponse_count=stats.get('empty_firstresponse_count', 0),
                daily_new_count=total_new,
                daily_closed_count=total_closed,
                upload_id=upload_record.id
            )
            db.session.add(statistic_record)
            db.session.commit()
            
            # Prepare response data
            response_data = {
                'success': True,
                'total_records': len(df),
                'new_records_count': new_records_count,
                'stats': stats,
                'filename': file.filename
            }
            
            return jsonify(response_data)
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': f'Error processing file: {str(e)}'}), 500
    
    return jsonify({'error': 'Invalid file format. Please upload Excel file (.xlsx or .xls)'}), 400

@app.route('/export/excel', methods=['POST'])
def export_excel():
    """Export analysis results to Excel with histogram"""
    try:
        data = request.get_json()
        if not data or 'stats' not in data:
            return jsonify({'error': 'No data to export'}), 400
        
        stats = data['stats']
        
        # Create Excel file in memory
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Summary sheet
            summary_data = {
                'Metric': ['Total Records', 'Current Open Tickets', 'Empty FirstResponse'],
                'Value': [
                    data.get('total_records', 0),
                    stats.get('current_open_count', 0),
                    stats.get('empty_firstresponse_count', 0)
                ]
            }
            pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)
            
            # Daily statistics sheet
            if 'daily_new' in stats and 'daily_closed' in stats:
                daily_data = []
                # Sort dates in ascending order for calculation (starting from earliest)
                all_dates_asc = sorted(set(stats['daily_new'].keys()) | set(stats['daily_closed'].keys()))
                
                # Calculate cumulative Open Tickets in chronological order (starting from earliest date)
                cumulative_open = 0
                daily_open_calculated = {}
                for date in all_dates_asc:
                    new_count = stats['daily_new'].get(date, 0)
                    closed_count = stats['daily_closed'].get(date, 0)
                    cumulative_open = cumulative_open + new_count - closed_count
                    daily_open_calculated[date] = cumulative_open
                
                # Sort dates in descending order for output (latest date first)
                all_dates_desc = sorted(set(stats['daily_new'].keys()) | set(stats['daily_closed'].keys()), reverse=True)
                
                for date in all_dates_desc:
                    new_count = stats['daily_new'].get(date, 0)
                    closed_count = stats['daily_closed'].get(date, 0)
                    open_count = daily_open_calculated.get(date, 0)
                    
                    daily_data.append({
                        'Date': date,
                        'New Tickets': new_count,
                        'Closed Tickets': closed_count,
                        'Open Tickets': open_count
                    })
                
                pd.DataFrame(daily_data).to_excel(writer, sheet_name='Daily Statistics', index=False)
            
            # Priority distribution
            if 'priority_distribution' in stats:
                priority_data = [{'Priority': k, 'Count': v} for k, v in stats['priority_distribution'].items()]
                pd.DataFrame(priority_data).to_excel(writer, sheet_name='Priority Distribution', index=False)
            
            # State distribution
            if 'state_distribution' in stats:
                state_data = [{'State': k, 'Count': v} for k, v in stats['state_distribution'].items()]
                pd.DataFrame(state_data).to_excel(writer, sheet_name='State Distribution', index=False)
            
            # Age segments
            if 'age_segments' in stats:
                age_data = [
                    {'Age Segment': '≤24 hours', 'Count': stats['age_segments']['age_24h']},
                    {'Age Segment': '24-48 hours', 'Count': stats['age_segments']['age_24_48h']},
                    {'Age Segment': '48-72 hours', 'Count': stats['age_segments']['age_48_72h']},
                    {'Age Segment': '>72 hours', 'Count': stats['age_segments']['age_72h']}
                ]
                pd.DataFrame(age_data).to_excel(writer, sheet_name='Age Segments', index=False)
            
            # Add age details and empty first response details
            # Get all tickets from database
            tickets = OtrsTicket.query.all()
            
            if tickets:
                # Convert tickets to DataFrame
                ticket_data = []
                for ticket in tickets:
                    ticket_data.append({
                        'TicketNumber': ticket.ticket_number,
                        'Created': ticket.created_date,
                        'Closed': ticket.closed_date,
                        'State': ticket.state,
                        'Priority': ticket.priority,
                        'FirstResponse': ticket.first_response,
                        'Age': ticket.age,
                        'AgeHours': ticket.age_hours
                    })
                
                df = pd.DataFrame(ticket_data)
                
                # Find actual column names
                possible_columns = {
                    'ticket_number': ['Ticket Number', 'TicketNumber', 'Number', 'ticket_number', 'id'],
                    'age': ['Age', 'age'],
                    'created': ['Created', 'CreateTime', 'Create Time', 'Date Created', 'created', 'creation_date'],
                    'priority': ['Priority', 'priority'],
                    'closed': ['Closed', 'CloseTime', 'Close Time', 'Date Closed', 'closed', 'close_date'],
                    'firstresponse': ['FirstResponse', 'First Response', 'firstresponse'],
                    'state': ['State', 'Status', 'Ticket State', 'state', 'status']
                }
                
                actual_columns = {}
                for key, possible_names in possible_columns.items():
                    for col in df.columns:
                        if any(name.lower() in col.lower() for name in possible_names):
                            actual_columns[key] = col
                            break
                
                # Age details sheets
                if 'age' in actual_columns and 'closed' in actual_columns:
                    open_tickets = df[df[actual_columns['closed']].isna()]
                    open_tickets = open_tickets.copy()
                    open_tickets['age_hours'] = open_tickets[actual_columns['age']].apply(parse_age_to_hours)
                    
                    # Age segments details
                    age_segments_details = {
                        '24h': open_tickets[open_tickets['age_hours'] <= 24],
                        '24_48h': open_tickets[(open_tickets['age_hours'] > 24) & (open_tickets['age_hours'] <= 48)],
                        '48_72h': open_tickets[(open_tickets['age_hours'] > 48) & (open_tickets['age_hours'] <= 72)],
                        '72h': open_tickets[open_tickets['age_hours'] > 72]
                    }
                    
                    for segment_name, segment_data in age_segments_details.items():
                        if not segment_data.empty:
                            segment_details = segment_data[[
                                actual_columns.get('ticket_number', 'Ticket Number'),
                                actual_columns.get('age', 'Age'),
                                actual_columns.get('created', 'Created'),
                                actual_columns.get('priority', 'Priority')
                            ]].copy()
                            segment_details.columns = ['Ticket Number', 'Age', 'Created', 'Priority']
                            sheet_name = f"Age {segment_name.replace('_', '-')} Details"
                            pd.DataFrame(segment_details).to_excel(writer, sheet_name=sheet_name[:31], index=False)
                
                # Empty first response details
                if 'firstresponse' in actual_columns:
                    firstresponse_col = actual_columns['firstresponse']
                    nan_empty = df[firstresponse_col].isna()
                    empty_strings = df[firstresponse_col] == ''
                    nan_strings = df[firstresponse_col].astype(str).str.lower() == 'nan'
                    
                    # Exclude Closed and Resolved states
                    if 'state' in actual_columns:
                        not_closed_resolved = ~df[actual_columns['state']].isin(['Closed', 'Resolved'])
                        empty_firstresponse = df[(nan_empty | empty_strings | nan_strings) & not_closed_resolved]
                    else:
                        empty_firstresponse = df[nan_empty | empty_strings | nan_strings]
                    
                    if not empty_firstresponse.empty:
                        empty_details = empty_firstresponse[[
                            actual_columns.get('ticket_number', 'Ticket Number'),
                            actual_columns.get('age', 'Age'),
                            actual_columns.get('created', 'Created'),
                            actual_columns.get('priority', 'Priority')
                        ]].copy()
                        empty_details.columns = ['Ticket Number', 'Age', 'Created', 'Priority']
                        pd.DataFrame(empty_details).to_excel(writer, sheet_name='Empty FirstResponse Details', index=False)
        
        # Generate histogram if daily data exists
        if 'daily_new' in stats and 'daily_closed' in stats:
            img_buffer = generate_histogram(stats['daily_new'], stats['daily_closed'])
            
            # Add histogram to Excel
            from openpyxl import load_workbook
            from openpyxl.drawing.image import Image
            
            output.seek(0)
            wb = load_workbook(output)
            
            # Create new sheet for histogram
            if 'Histogram' in wb.sheetnames:
                ws_hist = wb['Histogram']
            else:
                ws_hist = wb.create_sheet('Histogram')
            
            # Add image
            img = Image(img_buffer)
            img.width = 600
            img.height = 300
            ws_hist.add_image(img, 'A1')
            
            output = io.BytesIO()
            wb.save(output)
        
        output.seek(0)
        filename = f"otrs_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Record statistical query results to statistic table
        statistic_record = Statistic(
            query_type='export_excel',
            record_count=1  # Export operation recorded as 1 time
        )
        db.session.add(statistic_record)
        db.session.commit()
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': f'Error exporting Excel: {str(e)}'}), 500

@app.route('/age-details', methods=['POST'])
def get_age_details():
    """Get age segment details directly from database"""
    try:
        data = request.get_json()
        if not data or 'age_segment' not in data:
            return jsonify({'error': 'Missing required data'}), 400
        
        age_segment = data['age_segment']
        
        # Get open tickets from database (where closed_date is NULL)
        open_tickets = OtrsTicket.query.filter(OtrsTicket.closed_date.is_(None)).all()
        
        if not open_tickets:
            return jsonify({'error': 'No open tickets found in database'}), 400
        
        # Filter by age segment directly from database
        details = []
        for ticket in open_tickets:
            if ticket.age_hours is not None:
                if age_segment == '24h' and ticket.age_hours <= 24:
                    detail = {
                        'ticket_number': ticket.ticket_number or 'N/A',
                        'age': ticket.age or 'N/A',
                        'created': str(ticket.created_date) if ticket.created_date else 'N/A',
                        'priority': ticket.priority or 'N/A'
                    }
                    details.append(detail)
                elif age_segment == '24_48h' and 24 < ticket.age_hours <= 48:
                    detail = {
                        'ticket_number': ticket.ticket_number or 'N/A',
                        'age': ticket.age or 'N/A',
                        'created': str(ticket.created_date) if ticket.created_date else 'N/A',
                        'priority': ticket.priority or 'N/A'
                    }
                    details.append(detail)
                elif age_segment == '48_72h' and 48 < ticket.age_hours <= 72:
                    detail = {
                        'ticket_number': ticket.ticket_number or 'N/A',
                        'age': ticket.age or 'N/A',
                        'created': str(ticket.created_date) if ticket.created_date else 'N/A',
                        'priority': ticket.priority or 'N/A'
                    }
                    details.append(detail)
                elif age_segment == '72h' and ticket.age_hours > 72:
                    detail = {
                        'ticket_number': ticket.ticket_number or 'N/A',
                        'age': ticket.age or 'N/A',
                        'created': str(ticket.created_date) if ticket.created_date else 'N/A',
                        'priority': ticket.priority or 'N/A'
                    }
                    details.append(detail)
        
        # Record statistical query results to statistic table
        statistic_record = Statistic(
            query_type='age_details',
            age_segment=age_segment,
            record_count=len(details)
        )
        db.session.add(statistic_record)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'details': details
        })
        
    except Exception as e:
        return jsonify({'error': f'Error getting age details: {str(e)}'}), 500

@app.route('/empty-firstresponse-details', methods=['POST'])
def get_empty_firstresponse_details():
    """Get empty first response details directly from database"""
    try:
        # Get tickets with empty first response directly from database
        # Exclude Closed and Resolved states
        empty_firstresponse_tickets = OtrsTicket.query.filter(
            (OtrsTicket.first_response.is_(None) | 
             (OtrsTicket.first_response == '') |
             (OtrsTicket.first_response == 'nan') |
             (OtrsTicket.first_response == 'NaN')),
            ~OtrsTicket.state.in_(['Closed', 'Resolved'])
        ).all()
        
        # Prepare details
        details = []
        for ticket in empty_firstresponse_tickets:
            detail = {
                'ticket_number': ticket.ticket_number or 'N/A',
                'age': ticket.age or 'N/A',
                'created': str(ticket.created_date) if ticket.created_date else 'N/A',
                'priority': ticket.priority or 'N/A',
                'state': ticket.state or 'N/A'
            }
            details.append(detail)
        
        # Record statistical query results to statistic table
        statistic_record = Statistic(
            query_type='empty_firstresponse',
            record_count=len(details)
        )
        db.session.add(statistic_record)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'details': details
        })
        
    except Exception as e:
        return jsonify({'error': f'Error getting empty first response details: {str(e)}'}), 500

@app.route('/database-stats')
def database_stats():
    """Get comprehensive statistics directly from database"""
    try:
        # Get total records count
        total_records = OtrsTicket.query.count()
        
        if total_records == 0:
            return jsonify({
                'success': True,
                'total_records': 0,
                'data_sources_count': 0,
                'last_updated': None,
                'stats': {},
                'empty_firstresponse_details': []
            })
        
        # Get data sources count
        data_sources_count = OtrsTicket.query.with_entities(OtrsTicket.data_source).distinct().count()
        
        # Get last updated timestamp
        last_updated_ticket = OtrsTicket.query.order_by(OtrsTicket.import_time.desc()).first()
        last_updated = last_updated_ticket.import_time.isoformat() if last_updated_ticket and last_updated_ticket.import_time else None
        
        # Get statistics using direct database queries
        stats = analyze_otrs_tickets_direct_from_db()
        
        # Get empty first response details
        empty_firstresponse_tickets = OtrsTicket.query.filter(
            (OtrsTicket.first_response.is_(None) | 
             (OtrsTicket.first_response == '') |
             (OtrsTicket.first_response == 'nan') |
             (OtrsTicket.first_response == 'NaN')),
            ~OtrsTicket.state.in_(['Closed', 'Resolved'])
        ).all()
        
        empty_firstresponse_details = []
        for ticket in empty_firstresponse_tickets:
            detail = {
                'ticket_number': ticket.ticket_number or 'N/A',
                'age': ticket.age or 'N/A',
                'created': str(ticket.created_date) if ticket.created_date else 'N/A',
                'priority': ticket.priority or 'N/A',
                'state': ticket.state or 'N/A'
            }
            empty_firstresponse_details.append(detail)
        
        return jsonify({
            'success': True,
            'total_records': total_records,
            'data_sources_count': data_sources_count,
            'last_updated': last_updated,
            'stats': stats,
            'empty_firstresponse_details': empty_firstresponse_details
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Error getting database statistics: {str(e)}'
        }), 500

@app.route('/database')
def database_page():
    """Database statistics page"""
    return render_template('database_stats.html', APP_VERSION=APP_VERSION)

@app.route('/export/txt', methods=['POST'])
def export_txt():
    """Export analysis results to text file"""
    try:
        data = request.get_json()
        if not data or 'stats' not in data:
            return jsonify({'error': 'No data to export'}), 400
        
        stats = data['stats']
        
        # Create text content
        content = []
        content.append("=" * 60)
        content.append("OTRS TICKET ANALYSIS REPORT")
        content.append("=" * 60)
        content.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        content.append(f"Total Records: {data.get('total_records', 0)}")
        content.append("")
        
        # Summary
        content.append("SUMMARY")
        content.append("-" * 30)
        content.append(f"Current Open Tickets: {stats.get('current_open_count', 0)}")
        content.append(f"Empty FirstResponse: {stats.get('empty_firstresponse_count', 0)}")
        content.append("")
        
        # Daily statistics
        if 'daily_new' in stats and 'daily_closed' in stats:
            content.append("DAILY STATISTICS")
            content.append("-" * 30)
            content.append("Date\t\tNew\tClosed\tOpen")
            content.append("-" * 30)
            
            # Sort dates in ascending order for calculation (starting from earliest)
            all_dates_asc = sorted(set(stats['daily_new'].keys()) | set(stats['daily_closed'].keys()))
            
            # Calculate cumulative Open Tickets in chronological order (starting from earliest date)
            cumulative_open = 0
            daily_open_calculated = {}
            for date in all_dates_asc:
                new_count = stats['daily_new'].get(date, 0)
                closed_count = stats['daily_closed'].get(date, 0)
                cumulative_open = cumulative_open + new_count - closed_count
                daily_open_calculated[date] = cumulative_open
            
            # Sort dates in descending order for output (latest date first)
            all_dates_desc = sorted(set(stats['daily_new'].keys()) | set(stats['daily_closed'].keys()), reverse=True)
            
            for date in all_dates_desc:
                new_count = stats['daily_new'].get(date, 0)
                closed_count = stats['daily_closed'].get(date, 0)
                open_count = daily_open_calculated.get(date, 0)
                content.append(f"{date}\t{new_count}\t{closed_count}\t{open_count}")
            content.append("")
        
        # Priority distribution
        if 'priority_distribution' in stats:
            content.append("PRIORITY DISTRIBUTION")
            content.append("-" * 30)
            for priority, count in stats['priority_distribution'].items():
                content.append(f"{priority}: {count}")
            content.append("")
        
        # State distribution
        if 'state_distribution' in stats:
            content.append("STATE DISTRIBUTION")
            content.append("-" * 30)
            for state, count in stats['state_distribution'].items():
                content.append(f"{state}: {count}")
            content.append("")
        
        # Empty FirstResponse by priority
        if 'empty_firstresponse_by_priority' in stats:
            content.append("EMPTY FIRSTRESPONSE BY PRIORITY")
            content.append("-" * 30)
            for priority, count in stats['empty_firstresponse_by_priority'].items():
                content.append(f"{priority}: {count}")
            content.append("")
        
        # Age segments
        if 'age_segments' in stats:
            content.append("OPEN TICKETS AGE SEGMENT STATISTICS")
            content.append("-" * 30)
            content.append(f"≤24 hours: {stats['age_segments']['age_24h']}")
            content.append(f"24-48 hours: {stats['age_segments']['age_24_48h']}")
            content.append(f"48-72 hours: {stats['age_segments']['age_48_72h']}")
            content.append(f">72 hours: {stats['age_segments']['age_72h']}")
            content.append("")
        
        # Add age details and empty first response details
        # Get all tickets from database
        tickets = OtrsTicket.query.all()
        
        if tickets:
            # Convert tickets to DataFrame
            ticket_data = []
            for ticket in tickets:
                ticket_data.append({
                    'TicketNumber': ticket.ticket_number,
                    'Created': ticket.created_date,
                    'Closed': ticket.closed_date,
                    'State': ticket.state,
                    'Priority': ticket.priority,
                    'FirstResponse': ticket.first_response,
                    'Age': ticket.age,
                    'AgeHours': ticket.age_hours
                })
            
            df = pd.DataFrame(ticket_data)
            
            # Find actual column names
            possible_columns = {
                'ticket_number': ['Ticket Number', 'TicketNumber', 'Number', 'ticket_number', 'id'],
                'age': ['Age', 'age'],
                'created': ['Created', 'CreateTime', 'Create Time', 'Date Created', 'created', 'creation_date'],
                'priority': ['Priority', 'priority'],
                'closed': ['Closed', 'CloseTime', 'Close Time', 'Date Closed', 'closed', 'close_date'],
                'firstresponse': ['FirstResponse', 'First Response', 'firstresponse'],
                'state': ['State', 'Status', 'Ticket State', 'state', 'status']
            }
            
            actual_columns = {}
            for key, possible_names in possible_columns.items():
                for col in df.columns:
                    if any(name.lower() in col.lower() for name in possible_names):
                        actual_columns[key] = col
                        break
            
            # Age details
            if 'age' in actual_columns and 'closed' in actual_columns:
                open_tickets = df[df[actual_columns['closed']].isna()]
                open_tickets = open_tickets.copy()
                open_tickets['age_hours'] = open_tickets[actual_columns['age']].apply(parse_age_to_hours)
                
                # Age segments details
                age_segments_details = {
                    '≤24 hours': open_tickets[open_tickets['age_hours'] <= 24],
                    '24-48 hours': open_tickets[(open_tickets['age_hours'] > 24) & (open_tickets['age_hours'] <= 48)],
                    '48-72 hours': open_tickets[(open_tickets['age_hours'] > 48) & (open_tickets['age_hours'] <= 72)],
                    '>72 hours': open_tickets[open_tickets['age_hours'] > 72]
                }
                
                for segment_name, segment_data in age_segments_details.items():
                    if not segment_data.empty:
                        content.append(f"{segment_name} Ticket Details")
                        content.append("-" * 30)
                        content.append("Ticket Number\tAge\tCreated\tPriority")
                        content.append("-" * 30)
                        
                        for _, ticket in segment_data.iterrows():
                            ticket_number = str(ticket[actual_columns.get('ticket_number', 'Ticket Number')]) if 'ticket_number' in actual_columns else 'N/A'
                            age = str(ticket[actual_columns.get('age', 'Age')]) if 'age' in actual_columns else 'N/A'
                            created = str(ticket[actual_columns.get('created', 'Created')]) if 'created' in actual_columns else 'N/A'
                            priority = str(ticket[actual_columns.get('priority', 'Priority')]) if 'priority' in actual_columns else 'N/A'
                            content.append(f"{ticket_number}\t{age}\t{created}\t{priority}")
                        content.append("")
            
                # Empty first response details
                if 'firstresponse' in actual_columns:
                    firstresponse_col = actual_columns['firstresponse']
                    nan_empty = df[firstresponse_col].isna()
                    empty_strings = df[firstresponse_col] == ''
                    nan_strings = df[firstresponse_col].astype(str).str.lower() == 'nan'
                    
                    # Exclude Closed and Resolved states
                    if 'state' in actual_columns:
                        not_closed_resolved = ~df[actual_columns['state']].isin(['Closed', 'Resolved'])
                        empty_firstresponse = df[(nan_empty | empty_strings | nan_strings) & not_closed_resolved]
                    else:
                        empty_firstresponse = df[nan_empty | empty_strings | nan_strings]
                    
                    if not empty_firstresponse.empty:
                        content.append("Empty FirstResponse Ticket Details")
                        content.append("-" * 30)
                        content.append("Ticket Number\tAge\tCreated\tPriority")
                        content.append("-" * 30)
                        
                        for _, ticket in empty_firstresponse.iterrows():
                            ticket_number = str(ticket[actual_columns.get('ticket_number', 'Ticket Number')]) if 'ticket_number' in actual_columns else 'N/A'
                            age = str(ticket[actual_columns.get('age', 'Age')]) if 'age' in actual_columns else 'N/A'
                            created = str(ticket[actual_columns.get('created', 'Created')]) if 'created' in actual_columns else 'N/A'
                            priority = str(ticket[actual_columns.get('priority', 'Priority')]) if 'priority' in actual_columns else 'N/A'
                            content.append(f"{ticket_number}\t{age}\t{created}\t{priority}")
                        content.append("")
        
        # Convert to text file
        text_content = "\n".join(content)
        text_buffer = io.BytesIO()
        text_buffer.write(text_content.encode('utf-8'))
        text_buffer.seek(0)
        
        filename = f"otrs_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        
        # Record statistical query results to statistic table
        statistic_record = Statistic(
            query_type='export_txt',
            record_count=1  # Export operation recorded as 1 time
        )
        db.session.add(statistic_record)
        db.session.commit()
        
        return send_file(
            text_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='text/plain'
        )
        
    except Exception as e:
        return jsonify({'error': f'Error exporting text: {str(e)}'}), 500

@app.route('/clear-database', methods=['POST'])
def clear_database():
    """Clear all data from OTRS tickets table"""
    try:
        # Get current record count
        record_count = OtrsTicket.query.count()
        
        if record_count == 0:
            return jsonify({
                'success': True,
                'message': 'Database is already empty, no need to clear',
                'records_cleared': 0
            })
        
        # Delete all records
        OtrsTicket.query.delete()
        db.session.commit()
        
        # Log database operation
        log_database_operation(
            operation_type='clear_tickets',
            table_name='otrs_ticket',
            records_affected=record_count,
            operation_details='User manually cleared all ticket data',
            filename='manual_clear'
        )
        
        return jsonify({
            'success': True,
            'message': f'Successfully cleared {record_count} ticket records',
            'records_cleared': record_count
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': f'Error clearing database: {str(e)}'
        }), 500

@app.route('/progress')
def get_progress():
    """Get current processing progress status"""
    return jsonify(processing_status)

# Responsible Statistics APIs
@app.route('/responsible-stats')
def responsible_stats_page():
    """Responsible statistics page"""
    return render_template('responsible_stats.html', APP_VERSION=APP_VERSION)

@app.route('/api/responsible-list')
def get_responsible_list():
    """Get all unique Responsible values from database"""
    try:
        # Get all unique Responsible values
        responsibles = db.session.query(OtrsTicket.responsible).distinct().all()
        responsible_list = [r[0] for r in responsibles if r[0] is not None and r[0] != '']
        
        return jsonify({
            'success': True,
            'responsibles': sorted(responsible_list)
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Error getting responsible list: {str(e)}'
        }), 500

@app.route('/api/responsible-config', methods=['GET'])
def get_responsible_config():
    """Get user's Responsible configuration"""
    try:
        user_ip = request.remote_addr
        config = ResponsibleConfig.query.filter_by(user_identifier=user_ip).first()
        
        if config:
            import json
            selected_responsibles = json.loads(config.selected_responsibles)
        else:
            selected_responsibles = []
            
        return jsonify({
            'success': True,
            'selectedResponsibles': selected_responsibles
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Error getting responsible config: {str(e)}'
        }), 500

@app.route('/api/responsible-config', methods=['POST'])
def save_responsible_config():
    """Save user's Responsible configuration"""
    try:
        data = request.get_json()
        selected_responsibles = data.get('selectedResponsibles', [])
        user_ip = request.remote_addr
        
        import json
        
        # Check if config already exists
        config = ResponsibleConfig.query.filter_by(user_identifier=user_ip).first()
        
        if config:
            # Update existing config
            config.selected_responsibles = json.dumps(selected_responsibles)
            config.updated_at = datetime.utcnow()
        else:
            # Create new config
            config = ResponsibleConfig(
                user_identifier=user_ip,
                selected_responsibles=json.dumps(selected_responsibles)
            )
            db.session.add(config)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Configuration saved successfully'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': f'Error saving responsible config: {str(e)}'
        }), 500

@app.route('/api/responsible-stats', methods=['POST'])
def get_responsible_stats():
    """Get statistics for selected Responsible persons"""
    try:
        data = request.get_json()
        period = data.get('period', 'day')  # day/week/month
        selected_responsibles = data.get('selectedResponsibles', [])
        
        if not selected_responsibles:
            return jsonify({
                'success': True,
                'stats': {},
                'message': 'No responsible persons selected'
            })
        
        # Build query based on period - use closed_date for workload calculation
        if period == 'day':
            # Group by responsible and closed date
            stats_query = db.session.query(
                OtrsTicket.responsible,
                db.func.date(OtrsTicket.closed_date).label('period'),
                db.func.count(OtrsTicket.id).label('count')
            ).filter(
                OtrsTicket.responsible.in_(selected_responsibles),
                OtrsTicket.closed_date.isnot(None)  # Only count closed tickets
            ).group_by(
                OtrsTicket.responsible,
                db.func.date(OtrsTicket.closed_date)
            ).order_by(
                OtrsTicket.responsible,
                db.func.date(OtrsTicket.closed_date)
            )
            
        elif period == 'week':
            # Group by responsible and closed week
            stats_query = db.session.query(
                OtrsTicket.responsible,
                db.func.strftime('%Y-%W', OtrsTicket.closed_date).label('period'),
                db.func.count(OtrsTicket.id).label('count')
            ).filter(
                OtrsTicket.responsible.in_(selected_responsibles),
                OtrsTicket.closed_date.isnot(None)  # Only count closed tickets
            ).group_by(
                OtrsTicket.responsible,
                db.func.strftime('%Y-%W', OtrsTicket.closed_date)
            ).order_by(
                OtrsTicket.responsible,
                db.func.strftime('%Y-%W', OtrsTicket.closed_date)
            )
            
        elif period == 'month':
            # Group by responsible and closed month
            stats_query = db.session.query(
                OtrsTicket.responsible,
                db.func.strftime('%Y-%m', OtrsTicket.closed_date).label('period'),
                db.func.count(OtrsTicket.id).label('count')
            ).filter(
                OtrsTicket.responsible.in_(selected_responsibles),
                OtrsTicket.closed_date.isnot(None)  # Only count closed tickets
            ).group_by(
                OtrsTicket.responsible,
                db.func.strftime('%Y-%m', OtrsTicket.closed_date)
            ).order_by(
                OtrsTicket.responsible,
                db.func.strftime('%Y-%m', OtrsTicket.closed_date)
            )
        
        else:
            return jsonify({
                'success': False,
                'error': 'Invalid period specified'
            }), 400
        
        # Execute query and format results
        results = stats_query.all()
        stats = {}
        
        for responsible, period, count in results:
            if responsible not in stats:
                stats[responsible] = {}
            stats[responsible][period] = count
        
        # Calculate totals for each responsible
        totals = {}
        for responsible in selected_responsibles:
            total_count = OtrsTicket.query.filter(
                OtrsTicket.responsible == responsible
            ).count()
            totals[responsible] = total_count
        
        return jsonify({
            'success': True,
            'stats': stats,
            'totals': totals,
            'period': period
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Error getting responsible stats: {str(e)}'
        }), 500

@app.route('/api/responsible-details', methods=['POST'])
def get_responsible_details():
    """Get ticket details for a specific responsible person and period"""
    try:
        data = request.get_json()
        responsible = data.get('responsible')
        period = data.get('period')
        time_value = data.get('timeValue')  # Specific date, week, or month
        
        if not responsible or not period or not time_value:
            return jsonify({
                'success': False,
                'error': 'Missing required parameters'
            }), 400
        
        # Build query based on period - use closed_date for workload calculation
        query = OtrsTicket.query.filter(OtrsTicket.responsible == responsible)
        
        if period == 'day':
            # Filter by specific closed date
            query = query.filter(db.func.date(OtrsTicket.closed_date) == time_value)
        elif period == 'week':
            # Filter by specific closed week (format: YYYY-WW)
            query = query.filter(db.func.strftime('%Y-%W', OtrsTicket.closed_date) == time_value)
        elif period == 'month':
            # Filter by specific closed month (format: YYYY-MM)
            query = query.filter(db.func.strftime('%Y-%m', OtrsTicket.closed_date) == time_value)
        else:
            return jsonify({
                'success': False,
                'error': 'Invalid period specified'
            }), 400
        
        # Get tickets
        tickets = query.order_by(OtrsTicket.created_date.desc()).all()
        
        # Prepare details
        details = []
        for ticket in tickets:
            detail = {
                'ticket_number': ticket.ticket_number or 'N/A',
                'created': ticket.created_date.isoformat() if ticket.created_date else 'N/A',
                'closed': ticket.closed_date.isoformat() if ticket.closed_date else 'N/A',
                'state': ticket.state or 'N/A',
                'priority': ticket.priority or 'N/A',
                'title': ticket.title or 'N/A',
                'queue': ticket.queue or 'N/A'
            }
            details.append(detail)
        
        return jsonify({
            'success': True,
            'details': details,
            'count': len(details)
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Error getting responsible details: {str(e)}'
        }), 500

@app.route('/api/export-responsible-excel', methods=['POST'])
def export_responsible_excel():
    """Export Responsible statistics to Excel"""
    try:
        data = request.get_json()
        period = data.get('period', 'day')
        selected_responsibles = data.get('selectedResponsibles', [])
        stats_data = data.get('statsData', {})
        totals_data = data.get('totalsData', {})
        
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        
        # Create Excel file in memory
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Summary sheet
            summary_data = []
            for responsible, total in totals_data.items():
                summary_data.append({
                    'Responsible': responsible,
                    'Total Tickets': total
                })
            
            pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)
            
            # Detailed statistics sheet
            detailed_data = []
            for responsible, periods in stats_data.items():
                for period_name, count in periods.items():
                    detailed_data.append({
                        'Responsible': responsible,
                        'Period': period_name,
                        'Count': count
                    })
            
            pd.DataFrame(detailed_data).to_excel(writer, sheet_name='Detailed Stats', index=False)
            
            # Configuration sheet
            config_data = [{
                'Export Time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'Period': period,
                'Selected Responsibles': ', '.join(selected_responsibles)
            }]
            pd.DataFrame(config_data).to_excel(writer, sheet_name='Configuration', index=False)
        
        output.seek(0)
        filename = f"responsible_stats_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Error exporting Excel: {str(e)}'
        }), 500

@app.route('/api/export-responsible-txt', methods=['POST'])
def export_responsible_txt():
    """Export Responsible statistics to Text"""
    try:
        data = request.get_json()
        period = data.get('period', 'day')
        selected_responsibles = data.get('selectedResponsibles', [])
        stats_data = data.get('statsData', {})
        totals_data = data.get('totalsData', {})
        
        # Create text content
        content = []
        content.append("=" * 60)
        content.append("RESPONSIBLE WORKLOAD STATISTICS REPORT")
        content.append("=" * 60)
        content.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        content.append(f"Period: {period}")
        content.append(f"Selected Responsibles: {', '.join(selected_responsibles)}")
        content.append("")
        
        # Summary section
        content.append("SUMMARY")
        content.append("-" * 30)
        content.append("Responsible\tTotal Tickets")
        content.append("-" * 30)
        for responsible, total in totals_data.items():
            content.append(f"{responsible}\t{total}")
        content.append("")
        
        # Detailed statistics
        content.append("DETAILED STATISTICS")
        content.append("-" * 30)
        content.append("Responsible\tPeriod\tCount")
        content.append("-" * 30)
        for responsible, periods in stats_data.items():
            for period_name, count in periods.items():
                content.append(f"{responsible}\t{period_name}\t{count}")
        content.append("")
        
        # Convert to text file
        text_content = "\n".join(content)
        text_buffer = io.BytesIO()
        text_buffer.write(text_content.encode('utf-8'))
        text_buffer.seek(0)
        
        filename = f"responsible_stats_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        
        return send_file(
            text_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='text/plain'
        )
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Error exporting text: {str(e)}'
        }), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
