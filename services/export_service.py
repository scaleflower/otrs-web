"""
Export service for handling data export operations
"""

import io
import pandas as pd
from datetime import datetime
import matplotlib.pyplot as plt
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend

from models import db, OtrsTicket, StatisticsLog
from utils import generate_filename, get_user_info, parse_age_to_hours

class ExportService:
    """Service for export operations"""
    
    def __init__(self):
        pass
    
    def export_to_excel(self, analysis_data):
        """Export analysis results to Excel with histogram"""
        try:
            if not analysis_data or 'stats' not in analysis_data:
                raise ValueError('No data to export')
            
            stats = analysis_data['stats']
            
            # Create Excel file in memory
            output = io.BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Summary sheet
                summary_data = {
                    'Metric': ['Total Records', 'Current Open Tickets', 'Empty FirstResponse'],
                    'Value': [
                        analysis_data.get('total_records', 0),
                        stats.get('current_open_count', 0),
                        stats.get('empty_firstresponse_count', 0)
                    ]
                }
                pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)
                
                # Daily statistics sheet
                if 'daily_new' in stats and 'daily_closed' in stats:
                    daily_data = self._prepare_daily_data(stats)
                    pd.DataFrame(daily_data).to_excel(writer, sheet_name='Daily Statistics', index=False)
                
                # Priority distribution
                if 'priority_distribution' in stats:
                    priority_data = [{'Priority': k, 'Count': v} for k, v in stats['priority_distribution'].items()]
                    pd.DataFrame(priority_data).to_excel(writer, sheet_name='Priority Distribution', index=False)
                
                # State distribution
                if 'state_distribution' in stats:
                    state_data = [{'State': k, 'Count': v} for k, v in stats['state_distribution'].items()]
                    pd.DataFrame(state_data).to_excel(writer, sheet_name='State Distribution', index=False)
                
                # Age segments
                if 'age_segments' in stats:
                    age_data = [
                        {'Age Segment': '≤24 hours', 'Count': stats['age_segments']['age_24h']},
                        {'Age Segment': '24-48 hours', 'Count': stats['age_segments']['age_24_48h']},
                        {'Age Segment': '48-72 hours', 'Count': stats['age_segments']['age_48_72h']},
                        {'Age Segment': '>72 hours', 'Count': stats['age_segments']['age_72h']}
                    ]
                    pd.DataFrame(age_data).to_excel(writer, sheet_name='Age Segments', index=False)
                
                # Add detailed data sheets
                self._add_detailed_sheets(writer)
            
            # Generate histogram if daily data exists
            if 'daily_new' in stats and 'daily_closed' in stats:
                img_buffer = self._generate_histogram(stats['daily_new'], stats['daily_closed'])
                
                # Add histogram to Excel
                from openpyxl import load_workbook
                from openpyxl.drawing.image import Image
                
                output.seek(0)
                wb = load_workbook(output)
                
                # Create new sheet for histogram
                if 'Histogram' in wb.sheetnames:
                    ws_hist = wb['Histogram']
                else:
                    ws_hist = wb.create_sheet('Histogram')
                
                # Add image
                img = Image(img_buffer)
                img.width = 600
                img.height = 300
                ws_hist.add_image(img, 'A1')
                
                output = io.BytesIO()
                wb.save(output)
            
            output.seek(0)
            
            # Log export operation
            from .analysis_service import AnalysisService
            analysis_service = AnalysisService()
            analysis_service.log_statistic_query('export_excel', record_count=1)
            
            return output, generate_filename('otrs_analysis', 'xlsx')
            
        except Exception as e:
            raise Exception(f'Error exporting Excel: {str(e)}')
    
    def export_to_text(self, analysis_data):
        """Export analysis results to text file"""
        try:
            if not analysis_data or 'stats' not in analysis_data:
                raise ValueError('No data to export')
            
            stats = analysis_data['stats']
            
            # Create text content
            content = []
            content.append("=" * 60)
            content.append("OTRS TICKET ANALYSIS REPORT")
            content.append("=" * 60)
            content.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            content.append("")
            
            # Summary
            content.append("SUMMARY")
            content.append("-" * 40)
            content.append(f"Total Records: {analysis_data.get('total_records', 0)}")
            content.append(f"Current Open Tickets: {stats.get('current_open_count', 0)}")
            content.append(f"Empty FirstResponse: {stats.get('empty_firstresponse_count', 0)}")
            content.append("")
            
            # Daily statistics
            if 'daily_new' in stats and 'daily_closed' in stats:
                content.append("DAILY STATISTICS")
                content.append("-" * 40)
                all_dates = sorted(set(stats['daily_new'].keys()) | set(stats['daily_closed'].keys()), reverse=True)
                for date in all_dates:
                    new_count = stats['daily_new'].get(date, 0)
                    closed_count = stats['daily_closed'].get(date, 0)
                    content.append(f"{date}: New={new_count}, Closed={closed_count}")
                content.append("")
            
            # Priority distribution
            if 'priority_distribution' in stats:
                content.append("PRIORITY DISTRIBUTION")
                content.append("-" * 40)
                for priority, count in stats['priority_distribution'].items():
                    content.append(f"{priority}: {count}")
                content.append("")
            
            # State distribution
            if 'state_distribution' in stats:
                content.append("STATE DISTRIBUTION")
                content.append("-" * 40)
                for state, count in stats['state_distribution'].items():
                    content.append(f"{state}: {count}")
                content.append("")
            
            # Age segments with details
            if 'age_segments' in stats:
                content.append("AGE SEGMENTS (Open Tickets)")
                content.append("-" * 40)
                content.append(f"≤24 hours: {stats['age_segments']['age_24h']}")
                content.append(f"24-48 hours: {stats['age_segments']['age_24_48h']}")
                content.append(f"48-72 hours: {stats['age_segments']['age_48_72h']}")
                content.append(f">72 hours: {stats['age_segments']['age_72h']}")
                content.append("")
                
                # Add age segment details
                self._add_age_segment_details_to_text(content)
            
            # Convert to text
            text_content = "\n".join(content)
            
            # Create text file in memory
            output = io.BytesIO()
            output.write(text_content.encode('utf-8'))
            output.seek(0)
            
            # Log export operation
            from .analysis_service import AnalysisService
            analysis_service = AnalysisService()
            analysis_service.log_statistic_query('export_txt', record_count=1)
            
            return output, generate_filename('otrs_analysis', 'txt')
            
        except Exception as e:
            raise Exception(f'Error exporting text file: {str(e)}')
    
    def export_execution_logs(self):
        """Export all execution logs to Excel"""
        try:
            # Get all execution logs
            logs = StatisticsLog.query.order_by(StatisticsLog.execution_time.desc()).all()

            # Create Excel file in memory
            output = io.BytesIO()

            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Create logs data
                logs_data = []
                for log in logs:
                    logs_data.append({
                        'Execution Time': log.execution_time.isoformat() if log.execution_time else '',
                        'Statistic Date': str(log.statistic_date) if log.statistic_date else '',
                        'Age <24h': log.age_24h,
                        'Age 24-48h': log.age_24_48h,
                        'Age 48-72h': log.age_48_72h,
                        'Age 72-96h': log.age_72_96h,
                        'Total Open': log.total_open,
                        'Status': log.status,
                        'Error Message': log.error_message or '',
                        'Created At': log.created_at.isoformat() if log.created_at else ''
                    })

                # Write to Excel
                pd.DataFrame(logs_data).to_excel(writer, sheet_name='Execution Logs', index=False)

            output.seek(0)
            
            return output, generate_filename('execution_logs_export', 'xlsx')

        except Exception as e:
            raise Exception(f'Error exporting execution logs: {str(e)}')
    
    def _prepare_daily_data(self, stats):
        """Prepare daily statistics data for Excel export"""
        daily_data = []
        # Sort dates in ascending order for calculation (starting from earliest)
        all_dates_asc = sorted(set(stats['daily_new'].keys()) | set(stats['daily_closed'].keys()))
        
        # Calculate cumulative Open Tickets in chronological order (starting from earliest date)
        cumulative_open = 0
        daily_open_calculated = {}
        for date in all_dates_asc:
            new_count = stats['daily_new'].get(date, 0)
            closed_count = stats['daily_closed'].get(date, 0)
            cumulative_open = cumulative_open + new_count - closed_count
            daily_open_calculated[date] = cumulative_open
        
        # Sort dates in descending order for output (latest date first)
        all_dates_desc = sorted(set(stats['daily_new'].keys()) | set(stats['daily_closed'].keys()), reverse=True)
        
        for date in all_dates_desc:
            new_count = stats['daily_new'].get(date, 0)
            closed_count = stats['daily_closed'].get(date, 0)
            open_count = daily_open_calculated.get(date, 0)
            
            daily_data.append({
                'Date': date,
                'New Tickets': new_count,
                'Closed Tickets': closed_count,
                'Open Tickets': open_count
            })
        
        return daily_data
    
    def _add_detailed_sheets(self, writer):
        """Add detailed data sheets to Excel export"""
        # Get all tickets from database
        tickets = OtrsTicket.query.all()
        
        if not tickets:
            return
        
        # Convert tickets to DataFrame
        ticket_data = []
        for ticket in tickets:
            ticket_data.append({
                'TicketNumber': ticket.ticket_number,
                'Created': ticket.created_date,
                'Closed': ticket.closed_date,
                'State': ticket.state,
                'Priority': ticket.priority,
                'FirstResponse': ticket.first_response,
                'Age': ticket.age,
                'AgeHours': ticket.age_hours
            })
        
        df = pd.DataFrame(ticket_data)
        
        # Age details sheets
        open_tickets = df[df['Closed'].isna()]
        if not open_tickets.empty:
            open_tickets = open_tickets.copy()
            open_tickets['age_hours'] = open_tickets['Age'].apply(parse_age_to_hours)
            
            # Age segments details
            age_segments_details = {
                '24h': open_tickets[open_tickets['age_hours'] <= 24],
                '24_48h': open_tickets[(open_tickets['age_hours'] > 24) & (open_tickets['age_hours'] <= 48)],
                '48_72h': open_tickets[(open_tickets['age_hours'] > 48) & (open_tickets['age_hours'] <= 72)],
                '72h': open_tickets[open_tickets['age_hours'] > 72]
            }
            
            for segment_name, segment_data in age_segments_details.items():
                if not segment_data.empty:
                    segment_details = segment_data[['TicketNumber', 'Age', 'Created', 'Priority', 'State']].copy()
                    sheet_name = f"Age {segment_name.replace('_', '-')} Details"
                    segment_details.to_excel(writer, sheet_name=sheet_name[:31], index=False)
        
        # Empty first response details
        empty_firstresponse = df[
            (df['FirstResponse'].isna() | 
             (df['FirstResponse'] == '') |
             (df['FirstResponse'].astype(str).str.lower() == 'nan')) &
            (~df['State'].isin(['Closed', 'Resolved']))
        ]
        
        if not empty_firstresponse.empty:
            empty_details = empty_firstresponse[['TicketNumber', 'Age', 'Created', 'Priority']].copy()
            empty_details.to_excel(writer, sheet_name='Empty FirstResponse Details', index=False)
    
    def export_responsible_stats_to_excel(self, period, selected_responsibles, stats_data, totals_data, export_type='summary'):
        """Export responsible statistics to Excel"""
        try:
            output = io.BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                if export_type == 'summary':
                    # Export summary table
                    self._export_responsible_summary_excel(writer, period, selected_responsibles, stats_data, totals_data)
                else:
                    # Export details for each responsible person
                    self._export_responsible_details_excel(writer, period, selected_responsibles, stats_data, totals_data)
            
            output.seek(0)
            
            # Log export operation
            from .analysis_service import AnalysisService
            analysis_service = AnalysisService()
            analysis_service.log_statistic_query('export_responsible_excel', record_count=len(selected_responsibles))
            
            export_type_suffix = 'summary' if export_type == 'summary' else 'details'
            filename = generate_filename(f'responsible_stats_{export_type_suffix}', 'xlsx')
            return output, filename
            
        except Exception as e:
            raise Exception(f'Error exporting responsible statistics Excel: {str(e)}')
    
    def export_responsible_stats_to_text(self, period, selected_responsibles, stats_data, totals_data, export_type='summary'):
        """Export responsible statistics to text"""
        try:
            content = []
            content.append("=" * 80)
            content.append("RESPONSIBLE WORKLOAD STATISTICS REPORT")
            content.append("=" * 80)
            content.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            content.append(f"Period: {self._get_period_label(period)}")
            content.append(f"Export Type: {'汇总数据' if export_type == 'summary' else '明细数据'}")
            content.append(f"Selected Responsibles: {', '.join(selected_responsibles)}")
            content.append("")
            
            if export_type == 'summary':
                # Export summary table
                self._export_responsible_summary_text(content, period, selected_responsibles, stats_data, totals_data)
            else:
                # Export details for each responsible person
                self._export_responsible_details_text(content, period, selected_responsibles, stats_data, totals_data)
            
            # Convert to text
            text_content = "\n".join(content)
            
            # Create text file in memory
            output = io.BytesIO()
            output.write(text_content.encode('utf-8'))
            output.seek(0)
            
            # Log export operation
            from .analysis_service import AnalysisService
            analysis_service = AnalysisService()
            analysis_service.log_statistic_query('export_responsible_txt', record_count=len(selected_responsibles))
            
            export_type_suffix = 'summary' if export_type == 'summary' else 'details'
            filename = generate_filename(f'responsible_stats_{export_type_suffix}', 'txt')
            return output, filename
            
        except Exception as e:
            raise Exception(f'Error exporting responsible statistics text: {str(e)}')
    
    def _export_responsible_summary_excel(self, writer, period, selected_responsibles, stats_data, totals_data):
        """Export summary table to Excel"""
        if period == 'total':
            # Total statistics - simple ranking table
            summary_data = []
            sorted_totals = sorted(totals_data.items(), key=lambda x: x[1], reverse=True)
            
            for idx, (responsible, total) in enumerate(sorted_totals):
                summary_data.append({
                    '排名': idx + 1,
                    'Responsible人员': responsible,
                    '处理工单总数': total
                })
            
            pd.DataFrame(summary_data).to_excel(writer, sheet_name='汇总统计', index=False)
        else:
            # Period statistics - row/column swapped table
            period_stats = stats_data.get('period_stats', {})
            all_periods = sorted(period_stats.keys(), reverse=True)
            all_responsibles = sorted(selected_responsibles)
            
            # Build the table data
            table_data = []
            period_label = self._get_period_label(period)
            
            # Header row data
            header_row = {period_label: 'Period'}
            for responsible in all_responsibles:
                header_row[responsible] = responsible
            header_row['总计'] = '总计'
            
            # Period rows
            for period_key in all_periods:
                row = {period_label: period_key}
                period_total = 0
                
                for responsible in all_responsibles:
                    count = period_stats.get(period_key, {}).get(responsible, 0)
                    row[responsible] = count
                    period_total += count
                
                row['总计'] = period_total
                table_data.append(row)
            
            # Total row
            total_row = {period_label: '总计'}
            grand_total = 0
            for responsible in all_responsibles:
                responsible_total = totals_data.get(responsible, 0)
                total_row[responsible] = responsible_total
                grand_total += responsible_total
            total_row['总计'] = grand_total
            table_data.append(total_row)
            
            pd.DataFrame(table_data).to_excel(writer, sheet_name='汇总统计', index=False)
    
    def _export_responsible_details_excel(self, writer, period, selected_responsibles, stats_data, totals_data):
        """Export details for each responsible person to Excel"""
        period_stats = stats_data.get('period_stats', {})
        
        # Create a sheet for each responsible person
        for responsible in selected_responsibles:
            sheet_name = f"{responsible[:20]}明细"  # Limit name length
            
            details_data = []
            period_label = self._get_period_label(period)
            
            if period == 'total':
                # For total statistics, show a summary
                details_data.append({
                    period_label: '总体统计',
                    '工单数量': totals_data.get(responsible, 0)
                })
            else:
                # Show period breakdown
                all_periods = sorted(period_stats.keys(), reverse=True)
                for period_key in all_periods:
                    count = period_stats.get(period_key, {}).get(responsible, 0)
                    if count > 0:  # Only show periods with tickets
                        details_data.append({
                            period_label: period_key,
                            '工单数量': count
                        })
                
                # Add total row
                details_data.append({
                    period_label: '总计',
                    '工单数量': totals_data.get(responsible, 0)
                })
            
            if details_data:
                pd.DataFrame(details_data).to_excel(writer, sheet_name=sheet_name, index=False)
    
    def _export_responsible_summary_text(self, content, period, selected_responsibles, stats_data, totals_data):
        """Export summary table to text"""
        if period == 'total':
            content.append("SUMMARY STATISTICS")
            content.append("-" * 50)
            sorted_totals = sorted(totals_data.items(), key=lambda x: x[1], reverse=True)
            
            content.append(f"{'排名':<4} {'Responsible人员':<20} {'处理工单总数':<10}")
            content.append("-" * 50)
            
            for idx, (responsible, total) in enumerate(sorted_totals):
                content.append(f"{idx+1:<4} {responsible:<20} {total:<10}")
        else:
            period_stats = stats_data.get('period_stats', {})
            all_periods = sorted(period_stats.keys(), reverse=True)
            all_responsibles = sorted(selected_responsibles)
            period_label = self._get_period_label(period)
            
            content.append("PERIOD BREAKDOWN STATISTICS")
            content.append("-" * 80)
            
            # Header
            header = f"{period_label:<15}"
            for responsible in all_responsibles:
                header += f"{responsible:<12}"
            header += f"{'总计':<10}"
            content.append(header)
            content.append("-" * 80)
            
            # Period rows
            for period_key in all_periods:
                row = f"{period_key:<15}"
                period_total = 0
                
                for responsible in all_responsibles:
                    count = period_stats.get(period_key, {}).get(responsible, 0)
                    row += f"{count:<12}"
                    period_total += count
                
                row += f"{period_total:<10}"
                content.append(row)
            
            # Total row
            content.append("-" * 80)
            total_row = f"{'总计':<15}"
            grand_total = 0
            for responsible in all_responsibles:
                responsible_total = totals_data.get(responsible, 0)
                total_row += f"{responsible_total:<12}"
                grand_total += responsible_total
            total_row += f"{grand_total:<10}"
            content.append(total_row)
    
    def _export_responsible_details_text(self, content, period, selected_responsibles, stats_data, totals_data):
        """Export details for each responsible person to text"""
        period_stats = stats_data.get('period_stats', {})
        period_label = self._get_period_label(period)
        
        for responsible in selected_responsibles:
            content.append("")
            content.append(f"DETAILS FOR: {responsible}")
            content.append("=" * 60)
            
            if period == 'total':
                content.append(f"总体统计: {totals_data.get(responsible, 0)} 工单")
            else:
                content.append(f"{'Period':<20} {'工单数量':<10}")
                content.append("-" * 30)
                
                all_periods = sorted(period_stats.keys(), reverse=True)
                for period_key in all_periods:
                    count = period_stats.get(period_key, {}).get(responsible, 0)
                    if count > 0:
                        content.append(f"{period_key:<20} {count:<10}")
                
                content.append("-" * 30)
                content.append(f"{'总计':<20} {totals_data.get(responsible, 0):<10}")
    
    def _add_age_segment_details_to_text(self, content):
        """Add age segment details to text export"""
        try:
            # Get all open tickets from database
            tickets = OtrsTicket.query.filter(OtrsTicket.closed_date.is_(None)).all()
            
            if not tickets:
                content.append("No open tickets found for age segment details.")
                content.append("")
                return
            
            # Convert tickets to DataFrame for easier processing
            ticket_data = []
            for ticket in tickets:
                ticket_data.append({
                    'TicketNumber': ticket.ticket_number,
                    'Created': ticket.created_date,
                    'State': ticket.state,
                    'Priority': ticket.priority,
                    'Age': ticket.age,
                    'AgeHours': ticket.age_hours
                })
            
            df = pd.DataFrame(ticket_data)
            
            if df.empty:
                content.append("No open tickets data available.")
                content.append("")
                return
            
            # Parse age hours using utility function
            df['age_hours'] = df['Age'].apply(parse_age_to_hours)
            
            # Define age segments
            age_segments = {
                '≤24 hours': df[df['age_hours'] <= 24],
                '24-48 hours': df[(df['age_hours'] > 24) & (df['age_hours'] <= 48)],
                '48-72 hours': df[(df['age_hours'] > 48) & (df['age_hours'] <= 72)],
                '>72 hours': df[df['age_hours'] > 72]
            }
            
            # Add details for each segment
            for segment_name, segment_data in age_segments.items():
                if not segment_data.empty:
                    content.append(f"{segment_name.upper()} DETAILS")
                    content.append("-" * 60)
                    content.append(f"{'Ticket Number':<20} {'Age':<15} {'Created':<20} {'Priority':<10} {'State':<15}")
                    content.append("-" * 85)
                    
                    for _, ticket in segment_data.iterrows():
                        ticket_num = str(ticket['TicketNumber'])[:19] if ticket['TicketNumber'] else 'N/A'
                        age = str(ticket['Age'])[:14] if ticket['Age'] else 'N/A'
                        created = str(ticket['Created'])[:19] if ticket['Created'] else 'N/A'
                        priority = str(ticket['Priority'])[:9] if ticket['Priority'] else 'N/A'
                        state = str(ticket['State'])[:14] if ticket['State'] else 'N/A'
                        
                        content.append(f"{ticket_num:<20} {age:<15} {created:<20} {priority:<10} {state:<15}")
                    
                    content.append("")
                    
        except Exception as e:
            content.append(f"Error generating age segment details: {str(e)}")
            content.append("")
    
    def _get_period_label(self, period):
        """Get period label for display"""
        labels = {
            'total': '周期',
            'day': '日期',
            'week': '周次',
            'month': '月份'
        }
        return labels.get(period, '周期')
    
    def _generate_histogram(self, daily_new, daily_closed, daily_open=None):
        """Generate histogram for daily ticket statistics"""
        try:
            fig, ax = plt.subplots(figsize=(12, 6))
            
            # Prepare data
            all_dates = sorted(set(daily_new.keys()) | set(daily_closed.keys()))
            new_counts = [daily_new.get(date, 0) for date in all_dates]
            closed_counts = [daily_closed.get(date, 0) for date in all_dates]
            
            # Calculate open counts if provided
            if daily_open:
                open_counts = [daily_open.get(date, 0) for date in all_dates]
                bar_width = 0.25
                x_pos = range(len(all_dates))
                
                ax.bar([x - bar_width for x in x_pos], new_counts, bar_width, label='New Tickets', alpha=0.8, color='#2ecc71')
                ax.bar(x_pos, closed_counts, bar_width, label='Closed Tickets', alpha=0.8, color='#e74c3c')
                ax.bar([x + bar_width for x in x_pos], open_counts, bar_width, label='Open Tickets', alpha=0.8, color='#3498db')
            else:
                bar_width = 0.4
                x_pos = range(len(all_dates))
                
                ax.bar([x - bar_width/2 for x in x_pos], new_counts, bar_width, label='New Tickets', alpha=0.8, color='#2ecc71')
                ax.bar([x + bar_width/2 for x in x_pos], closed_counts, bar_width, label='Closed Tickets', alpha=0.8, color='#e74c3c')
            
            ax.set_xlabel('Date')
            ax.set_ylabel('Number of Tickets')
            ax.set_title('Daily Ticket Statistics')
            ax.set_xticks(x_pos)
            ax.set_xticklabels([str(date) for date in all_dates], rotation=45, ha='right')
            ax.legend()
            ax.grid(True, alpha=0.3)
            
            plt.tight_layout()
            
            # Save to buffer
            img_buffer = io.BytesIO()
            plt.savefig(img_buffer, format='png', dpi=100, bbox_inches='tight')
            img_buffer.seek(0)
            plt.close()
            
            return img_buffer
        except Exception as e:
            print(f"Warning: Could not generate histogram: {str(e)}")
            return None
